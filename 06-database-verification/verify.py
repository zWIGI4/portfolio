# -*- coding: utf-8 -*-
"""Database verification: our internal database vs a fresh market snapshot.

Input:  input/internal_database.csv + input/external_source.csv
Output: output/verification_report.xlsx
  - Summary          how many match, how many differ, what to add, what to explain
  - Discrepancies    field by field: our value vs the value in the source
  - To add           developments present in the source, missing from our database
  - To explain       developments from our database absent from the source
  - Matching         paired records with no differences

Matching survives different spellings: normalisation (lowercase, Polish
diacritics stripped, "etap II" = "etap 2", dashes as spaces) plus fuzzy matching
(difflib) for typos. Nothing is deleted automatically: the report points at
differences, a human decides.
"""
import csv
import difflib
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = Path(__file__).parent
OUT = BASE / "output"
OUT.mkdir(exist_ok=True)

# ---------- palette ----------
BRAND = "2A78D6"
BRAND_D = "1C5CAB"
AQUA = "1BAF7A"
INK = "0B0B0B"
INK2 = "52514E"
MUTED = "898781"
TILE = "EEF4FC"
TILE_LN = "B7D3F6"
GOOD = "0CA30C"
BAD = "D03B3B"
BAD_BG = "FDECEC"

thin = Side(style="thin", color=TILE_LN)
TILE_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def page(ws, fit_height=1):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

def brand_bar(ws, to_column=13):
    for c in range(1, to_column + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 6

def sheet_title(ws, text, subtitle):
    page(ws)
    brand_bar(ws)
    ws.cell(row=3, column=2, value=text).font = Font(name="Calibri", bold=True, size=17, color=INK)
    ws.cell(row=4, column=2, value=subtitle).font = Font(name="Calibri", size=10, color=MUTED)

def tile(ws, row, col, label, value, color=INK):
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col + 2)
    for rr in range(row, row + 3):
        for cc in range(col, col + 3):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill("solid", fgColor=TILE)
            cell.border = TILE_BORDER
    lab = ws.cell(row=row - 1, column=col, value=label.upper())
    lab.font = Font(name="Calibri", size=8.5, color=MUTED, bold=True)
    val = ws.cell(row=row, column=col, value=value)
    val.font = Font(name="Calibri", bold=True, size=22, color=color)
    val.alignment = Alignment(horizontal="center", vertical="center")

def table(ws, ref, name, header_color=BRAND, style="TableStyleMedium2"):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True,
                                      showFirstColumn=False, showLastColumn=False)
    ws.add_table(t)
    first, last = ref.split(":")
    for c in range(ws[first].column, ws[last].column + 1):
        cell = ws.cell(row=ws[first].row, column=c)
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10.5)
        cell.alignment = Alignment(horizontal="left")

ROMAN = {"etap i": "etap 1", "etap ii": "etap 2", "etap iii": "etap 3"}

def norm(text):
    t = text.lower().replace("-", " ")
    t = "".join(c for c in unicodedata.normalize("NFKD", t) if not unicodedata.combining(c))
    t = " ".join(t.split())
    for roman, arabic in sorted(ROMAN.items(), key=lambda x: -len(x[0])):
        if t.endswith(roman):
            t = t[: -len(roman)] + arabic
    return t

def load(name):
    with open(BASE / "input" / name, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

internal = load("internal_database.csv")
external = load("external_source.csv")

# ---------- matching ----------
external_keys = {i: (norm(e["inwestycja"]), e["miasto"].lower()) for i, e in enumerate(external)}
used_external = set()
pairs = []            # (internal_record, external_record, method)
unmatched_internal = []

for rec in internal:
    key_int = (norm(rec["nazwa_inwestycji"]), rec["miasto"].lower())
    hit = None
    for i, key_ext in external_keys.items():
        if i in used_external:
            continue
        if key_ext == key_int:
            hit = (i, "exact")
            break
    if hit is None:  # fuzzy, within the same city
        best, score = None, 0.0
        for i, key_ext in external_keys.items():
            if i in used_external or key_ext[1] != key_int[1]:
                continue
            r = difflib.SequenceMatcher(None, key_int[0], key_ext[0]).ratio()
            if r > score:
                best, score = i, r
        if best is not None and score >= 0.85:
            hit = (best, f"fuzzy {score:.0%}")
    if hit:
        used_external.add(hit[0])
        pairs.append((rec, external[hit[0]], hit[1]))
    else:
        unmatched_internal.append(rec)

to_add = [e for i, e in enumerate(external) if i not in used_external]

# ---------- field comparison ----------
FIELDS = [  # (our field, source field, label, comparison function)
    ("dzielnica", "dzielnica", "District", lambda a, b: norm(a) == norm(b)),
    ("liczba_lokali", "lokale", "Units", lambda a, b: str(a) == str(b)),
    ("status", "status", "Status", lambda a, b: a == b),
    ("termin_oddania", "oddanie", "Completion", lambda a, b: a == b),
    ("cena_za_m2", "cena_m2", "Price per m2", lambda a, b: str(a) == str(b)),
]

discrepancies = []
matching = []
for rec, ext, method in pairs:
    differences = []
    for f_int, f_ext, label, equal in FIELDS:
        if not equal(rec[f_int], ext[f_ext]):
            differences.append((label, rec[f_int], ext[f_ext]))
    if differences:
        for label, ours, theirs in differences:
            discrepancies.append({
                "id": rec["id"], "development": rec["nazwa_inwestycji"],
                "city": rec["miasto"], "field": label,
                "ours": ours, "source": theirs, "match": method,
            })
    else:
        matching.append((rec, ext, method))

# ---------- writing the report ----------
def widths(ws, values):
    for i, v in enumerate(values, start=1):
        ws.column_dimensions[get_column_letter(i)].width = v

wb = Workbook()

# --- Summary: tiles like in a management report ---
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
sheet_title(ws, "Investment database verification report",
            f"Internal database ({len(internal)}) vs fresh market snapshot ({len(external)}) · fictional data (demo)")
fuzzy_n = sum(1 for *_, s in pairs if s != "exact")
tiles = [
    ("Matched pairs", len(pairs), INK),
    ("Fully matching", len(matching), GOOD),
    ("Discrepancies (fields)", len(discrepancies), BAD),
    ("Fuzzy matches", fuzzy_n, AQUA),
    ("To add on our side", len(to_add), BRAND_D),
    ("To explain", len(unmatched_internal), BRAND_D),
]
for i, (label, value, color) in enumerate(tiles):
    tile(ws, 7 + (i // 3) * 5, 2 + (i % 3) * 4, label, value, color)
ws.cell(row=17, column=2,
        value="Nothing was changed automatically. The report points at differences, decisions stay with a human.")
ws.cell(row=17, column=2).font = Font(name="Calibri", size=9, color=MUTED, italic=True)
widths(ws, [2.5, 12, 12, 12, 3, 12, 12, 12, 3, 12, 12, 12])

# --- Discrepancies ---
ws = wb.create_sheet("Discrepancies")
ws.sheet_view.showGridLines = False
sheet_title(ws, "Discrepancies field by field",
            "Each row is one difference between our database and the source")
start = 6
for c, label in enumerate(["ID", "Development", "City", "Field", "Ours", "Source", "Match"], start=2):
    ws.cell(row=start, column=c, value=label)
for i, r in enumerate(discrepancies, start=start + 1):
    for c, key in enumerate(["id", "development", "city", "field", "ours", "source", "match"], start=2):
        ws.cell(row=i, column=c, value=r[key])
    for c in (6, 7):  # the values that differ, on a light red background
        ws.cell(row=i, column=c).fill = PatternFill("solid", fgColor=BAD_BG)
table(ws, f"B{start}:H{start+len(discrepancies)}", "TabDiscrepancies",
      header_color=BAD, style="TableStyleLight10")
ws.freeze_panes = f"A{start+1}"
widths(ws, [2.5, 10, 28, 12, 16, 16, 16, 16])

# --- To add ---
ws = wb.create_sheet("To add")
ws.sheet_view.showGridLines = False
sheet_title(ws, "New items from the source", "Present in the market snapshot, missing from our database")
start = 6
for c, label in enumerate(["Development", "City", "District", "Units", "Status", "Completion", "Price/m2"], start=2):
    ws.cell(row=start, column=c, value=label)
for i, e in enumerate(to_add, start=start + 1):
    for c, key in enumerate(["inwestycja", "miasto", "dzielnica", "lokale", "status", "oddanie", "cena_m2"], start=2):
        ws.cell(row=i, column=c, value=e[key])
table(ws, f"B{start}:H{start+len(to_add)}", "TabToAdd", header_color=AQUA, style="TableStyleLight9")
widths(ws, [2.5, 28, 12, 16, 10, 14, 10, 10])

# --- To explain ---
ws = wb.create_sheet("To explain")
ws.sheet_view.showGridLines = False
sheet_title(ws, "Ours, missing from the source", "Withdrawn from the market or renamed — check manually")
start = 6
for c, label in enumerate(["ID", "Developer", "Development", "City", "Status", "Note"], start=2):
    ws.cell(row=start, column=c, value=label)
for i, rec in enumerate(unmatched_internal, start=start + 1):
    for c, v in enumerate([rec["id"], rec["deweloper"], rec["nazwa_inwestycji"], rec["miasto"], rec["status"],
                           "missing from the current snapshot: withdrawn or renamed"], start=2):
        ws.cell(row=i, column=c, value=v)
table(ws, f"B{start}:G{start+len(unmatched_internal)}", "TabToExplain")
widths(ws, [2.5, 10, 26, 28, 12, 14, 44])

# --- Matching ---
ws = wb.create_sheet("Matching")
ws.sheet_view.showGridLines = False
sheet_title(ws, "Fully matching pairs", "Paired with no differences in the compared fields")
start = 6
for c, label in enumerate(["ID", "Development", "City", "Match"], start=2):
    ws.cell(row=start, column=c, value=label)
for i, (rec, ext, method) in enumerate(matching, start=start + 1):
    for c, v in enumerate([rec["id"], rec["nazwa_inwestycji"], rec["miasto"], method], start=2):
        ws.cell(row=i, column=c, value=v)
table(ws, f"B{start}:E{start+len(matching)}", "TabMatching")
widths(ws, [2.5, 10, 28, 12, 16])

wb.save(OUT / "verification_report.xlsx")
print("OK -> output/verification_report.xlsx")
print(f"  pairs: {len(pairs)} (fuzzy: {sum(1 for *_, s in pairs if s != 'exact')}) "
      f"| matching: {len(matching)} | discrepancies: {len(discrepancies)} fields")
print(f"  to add: {len(to_add)} | to explain: {len(unmatched_internal)}")
