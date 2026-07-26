#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sales dashboard built from a raw CSV export.
Input:  input/sales.csv (raw transactions)
Output: output/dashboard.xlsx — sheets:
        • KPI (tiles: revenue, orders, average order, top category...)
        • By month (table with data bars + bar chart)
        • By category (table with % share + chart)
        • Data (raw records as a filterable Excel table)

Run:  python3 dashboard.py
The result is a single self-contained .xlsx file — nothing to install on the
client side.
"""
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = Path(__file__).resolve().parent
INPUT = BASE / "input" / "sales.csv"
OUTPUT = BASE / "output" / "dashboard.xlsx"

MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ---------- palette ----------
BRAND = "2A78D6"
BRAND_D = "1C5CAB"
AQUA = "1BAF7A"
INK = "0B0B0B"
INK2 = "52514E"
MUTED = "898781"
TILE = "EEF4FC"
TILE_LN = "B7D3F6"
LINE = "E1E0D9"

thin = Side(style="thin", color=TILE_LN)
TILE_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def page(ws, height=1):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = height
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

def brand_bar(ws, to_col=13):
    for c in range(1, to_col + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 6

def title(ws, text, subtitle):
    page(ws)
    brand_bar(ws)
    ws.cell(row=3, column=2, value=text).font = Font(name="Calibri", bold=True, size=17, color=INK)
    ws.cell(row=4, column=2, value=subtitle).font = Font(name="Calibri", size=10, color=MUTED)

def tile(ws, row, col, label, value, color=INK, number_format=None):
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col + 2)
    for rr in range(row, row + 3):
        for cc in range(col, col + 3):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill("solid", fgColor=TILE)
            cell.border = TILE_BORDER
    lbl = ws.cell(row=row - 1, column=col, value=label.upper())
    lbl.font = Font(name="Calibri", size=8.5, color=MUTED, bold=True)
    val = ws.cell(row=row, column=col, value=value)
    val.font = Font(name="Calibri", bold=True, size=20, color=color)
    val.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        val.number_format = number_format

def table(ws, ref, name, style="TableStyleMedium2", header_color=BRAND):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True,
                                      showFirstColumn=False, showLastColumn=False)
    ws.add_table(t)
    first, last = ref.split(":")
    for c in range(ws[first].column, ws[last].column + 1):
        cell = ws.cell(row=ws[first].row, column=c)
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10.5)
        cell.alignment = Alignment(horizontal="left")

def style_chart(chart, color=BRAND):
    chart.series[0].graphicalProperties.solidFill = color
    chart.series[0].graphicalProperties.line.noFill = True
    chart.gapWidth = 60
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=LINE))
    chart.y_axis.numFmt = "#,##0"
    chart.legend = None

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    with open(INPUT, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    # CSV column names stay as in the source export (Polish demo data)
    for r in rows:
        r["wartosc"] = float(r["wartosc"])
        r["ilosc"] = int(r["ilosc"])
        r["month"] = int(r["data"][5:7])

    revenue = sum(r["wartosc"] for r in rows)
    orders = len(rows)
    average = revenue / orders if orders else 0
    by_category = defaultdict(float)
    by_month = defaultdict(float)
    by_channel = defaultdict(float)
    for r in rows:
        by_category[r["kategoria"]] += r["wartosc"]
        by_month[r["month"]] += r["wartosc"]
        by_channel[r["kanal"]] += r["wartosc"]
    top_category = max(by_category, key=by_category.get)
    top_month = max(by_month, key=by_month.get)
    top_channel = max(by_channel, key=by_channel.get)

    wb = Workbook()

    # --- KPI ---
    ws = wb.active
    ws.title = "KPI"
    ws.sheet_view.showGridLines = False
    title(ws, "Sales dashboard 2025",
          "Built automatically from a raw transaction export · fictional data (demo)")
    tiles = [
        ("Total revenue", revenue, INK, '#,##0" zł"'),
        ("Orders", orders, INK, "0"),
        ("Average order value", average, BRAND_D, '#,##0" zł"'),
        ("Top category", top_category, BRAND_D, None),
        ("Best month", MONTHS[top_month], INK2, None),
        (f"Channel share: {top_channel}", by_channel[top_channel] / revenue, AQUA, "0%"),
    ]
    for i, (label, value, color, fmt) in enumerate(tiles):
        tile(ws, 7 + (i // 3) * 5, 2 + (i % 3) * 4, label, value, color, fmt)
    set_widths(ws, [2.5, 12, 12, 12, 3, 12, 12, 12, 3, 12, 12, 12])

    # --- By month ---
    ws = wb.create_sheet("By month")
    ws.sheet_view.showGridLines = False
    title(ws, "Revenue by month", "Sum of order values · zł")
    start = 6
    ws.cell(row=start, column=2, value="Month")
    ws.cell(row=start, column=3, value="Revenue")
    for m in range(1, 13):
        ws.cell(row=start + m, column=2, value=MONTHS[m])
        ws.cell(row=start + m, column=3, value=round(by_month.get(m, 0), 2)).number_format = '#,##0" zł"'
    table(ws, f"B{start}:C{start+12}", "TabMonths")
    ws.conditional_formatting.add(
        f"C{start+1}:C{start+12}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=BRAND, showValue=True))
    set_widths(ws, [2.5, 12, 16])
    chart = BarChart()
    chart.title = "Revenue by month (zł)"
    chart.height, chart.width = 10, 18
    chart.add_data(Reference(ws, min_col=3, min_row=start, max_row=start + 12), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=2, min_row=start + 1, max_row=start + 12))
    style_chart(chart)
    ws.add_chart(chart, "E6")

    # --- By category ---
    ws = wb.create_sheet("By category")
    ws.sheet_view.showGridLines = False
    title(ws, "Revenue by category", "Share of total revenue")
    start = 6
    for c, name in enumerate(["Category", "Revenue", "Share"], start=2):
        ws.cell(row=start, column=c, value=name)
    categories = sorted(by_category.items(), key=lambda x: -x[1])
    for i, (k, v) in enumerate(categories, start=start + 1):
        ws.cell(row=i, column=2, value=k)
        ws.cell(row=i, column=3, value=round(v, 2)).number_format = '#,##0" zł"'
        ws.cell(row=i, column=4, value=v / revenue).number_format = "0%"
    n = len(categories)
    table(ws, f"B{start}:D{start+n}", "TabCategories")
    ws.conditional_formatting.add(
        f"C{start+1}:C{start+n}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=AQUA, showValue=True))
    set_widths(ws, [2.5, 18, 16, 10])
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Revenue by category (zł)"
    chart2.height, chart2.width = 10, 16
    chart2.add_data(Reference(ws, min_col=3, min_row=start, max_row=start + n), titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=2, min_row=start + 1, max_row=start + n))
    style_chart(chart2, color=AQUA)
    ws.add_chart(chart2, "F6")

    # --- Raw data ---
    ws = wb.create_sheet("Data")
    columns = ["nr_zamowienia", "data", "kategoria", "kanal", "ilosc", "cena_jedn", "wartosc"]
    for c, name in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=name)
    for r in rows:
        ws.append([r[c] for c in columns])
    table(ws, f"A1:G{len(rows)+1}", "TabData")
    ws.freeze_panes = "A2"
    set_widths(ws, [16, 12, 14, 14, 8, 11, 11])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"OK: dashboard from {orders} transactions")
    print(f"   revenue {revenue:,.0f} zł | top category: {top_category}".replace(",", " "))
    print("   file: output/dashboard.xlsx (4 sheets, 2 charts, Excel tables)")


if __name__ == "__main__":
    main()
