#!/usr/bin/env python3
"""
Ekstrakcja danych z faktur PDF → rejestr Excel (demo „dostarcz-i-koniec").
Wejście:  dane_wejsciowe/*.pdf  (faktury)
Wyjście:  wyniki/rejestr_faktur.xlsx — 2 arkusze:
          • Faktury  (nr, daty, sprzedawca, NIP-y, netto, brutto)
          • Pozycje  (wszystkie pozycje ze wszystkich faktur, do filtrowania)

Uruchomienie:  python3 ekstrakcja.py
W realnym zleceniu dopasowuję parser do układu faktur klienta (różni wystawcy —
różne szablony); to demo pokazuje pełen przepływ na spójnym układzie.
"""
import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BAZA = Path(__file__).resolve().parent
WEJSCIE = BAZA / "dane_wejsciowe"
WYJSCIE = BAZA / "wyniki" / "rejestr_faktur.xlsx"

NAGL = PatternFill("solid", fgColor="1F3864")
NAGL_F = Font(bold=True, color="FFFFFF")


def kwota(s):
    """'1 234,56 zł' -> 1234.56"""
    s = s.replace("zł", "").replace("\xa0", " ").strip().replace(" ", "").replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def blok_stron(strona):
    """Nazwy sprzedawcy/nabywcy z dwukolumnowego bloku — po współrzędnych.

    extract_text() skleja kolumny, więc wycinamy (crop) lewą i prawą połowę
    obszaru między nagłówkiem „Sprzedawca" a początkiem tabeli pozycji („Lp.")
    i bierzemy pierwszą linię każdej połowy.
    """
    slowa = strona.extract_words()
    y_top = next((w["bottom"] for w in slowa if w["text"] == "Sprzedawca"), None)
    y_bot = next((w["top"] for w in slowa if w["text"].startswith("Lp")), strona.height)
    if y_top is None:
        return None, None
    srodek = strona.width / 2
    # within_bbox (nie crop): bierze tylko obiekty w całości wewnątrz obszaru,
    # +2pt marginesu, żeby nie łapać samego nagłówka „Sprzedawca"
    lewa = strona.within_bbox((0, y_top + 2, srodek, y_bot)).extract_text() or ""
    prawa = strona.within_bbox((srodek, y_top + 2, strona.width, y_bot)).extract_text() or ""
    nazwa = lambda t: next((ln.strip() for ln in t.splitlines() if ln.strip()), None)
    return nazwa(lewa), nazwa(prawa)


def parsuj(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        strona = pdf.pages[0]
        tekst = strona.extract_text() or ""
        tabele = strona.extract_tables()
        sprzedawca_xy, nabywca_xy = blok_stron(strona)

    dane = {"plik": pdf_path.name}
    m = re.search(r"Faktura VAT\s+(\S+)", tekst)
    dane["nr_faktury"] = m.group(1) if m else None
    m = re.search(r"Data wystawienia:\s*([\d-]+)", tekst)
    dane["data_wystawienia"] = m.group(1) if m else None
    m = re.search(r"Termin płatności:\s*([\d-]+)", tekst)
    dane["termin_platnosci"] = m.group(1) if m else None

    # sprzedawca/nabywca: z cropów po współrzędnych (nie z posklejanego tekstu)
    nipy = re.findall(r"NIP\s?(\d{10})", tekst)
    dane["nip_sprzedawcy"] = nipy[0] if len(nipy) > 0 else None
    dane["nip_nabywcy"] = nipy[1] if len(nipy) > 1 else None
    dane["sprzedawca"] = sprzedawca_xy
    dane["nabywca"] = nabywca_xy

    m = re.search(r"Do zapłaty:\s*([\d\s.,]+zł)\s*\(netto:\s*([\d\s.,]+zł)\)", tekst)
    dane["brutto"] = kwota(m.group(1)) if m else None
    dane["netto"] = kwota(m.group(2)) if m else None

    pozycje = []
    for tab in tabele:
        if not tab or not tab[0]:
            continue
        naglowek = [c or "" for c in tab[0]]
        if "Lp." not in naglowek[0]:
            continue
        for w in tab[1:]:
            if not w or not (w[0] or "").strip().isdigit():
                continue
            pozycje.append({
                "faktura": dane["nr_faktury"],
                "lp": int(w[0]),
                "nazwa": (w[1] or "").replace("\n", " ").strip(),
                "ilosc": kwota(w[2] or "") or (w[2] or "").strip(),
                "cena_netto": kwota(w[3] or ""),
                "vat": (w[4] or "").strip(),
                "brutto": kwota(w[5] or ""),
            })
    return dane, pozycje


def naglowek(ws, kolumny):
    ws.append(kolumny)
    for c in ws[1]:
        c.fill, c.font = NAGL, NAGL_F
    ws.freeze_panes = "A2"


def main():
    pliki = sorted(WEJSCIE.glob("*.pdf"))
    if not pliki:
        print("Brak PDF-ów w dane_wejsciowe/ — najpierw: python3 generuj_faktury.py")
        return
    faktury, pozycje = [], []
    for p in pliki:
        f, poz = parsuj(p)
        faktury.append(f)
        pozycje += poz
        print(f"OK {p.name}: {f['nr_faktury']} | brutto {f['brutto']} | pozycji {len(poz)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Faktury"
    kol_f = ["nr_faktury", "data_wystawienia", "termin_platnosci", "sprzedawca",
             "nip_sprzedawcy", "nabywca", "nip_nabywcy", "netto", "brutto", "plik"]
    naglowek(ws, kol_f)
    for f in faktury:
        ws.append([f.get(k) for k in kol_f])
    for i, w in enumerate([16, 15, 15, 30, 14, 22, 13, 11, 11, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wp = wb.create_sheet("Pozycje")
    kol_p = ["faktura", "lp", "nazwa", "ilosc", "cena_netto", "vat", "brutto"]
    naglowek(wp, kol_p)
    for poz in pozycje:
        wp.append([poz.get(k) for k in kol_p])
    wp.auto_filter.ref = f"A1:G{len(pozycje) + 1}"
    for i, w in enumerate([16, 6, 46, 8, 12, 8, 12], 1):
        wp.column_dimensions[get_column_letter(i)].width = w

    WYJSCIE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WYJSCIE)
    suma = sum(f["brutto"] or 0 for f in faktury)
    print(f"---\nRejestr: {len(faktury)} faktur, {len(pozycje)} pozycji, suma brutto {suma:,.2f} zł".replace(",", " "))
    print(f"Plik: {WYJSCIE.relative_to(BAZA)}")


if __name__ == "__main__":
    main()
