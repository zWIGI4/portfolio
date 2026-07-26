# -*- coding: utf-8 -*-
"""KPI dashboard from three systems: payments + bookings + attendance.

Input:   input/payments.csv, bookings.csv, attendance.csv
Output:  output/kpi_dashboard.xlsx
  - KPI             the key numbers in one place
  - By month        net revenue + bar chart
  - By course       revenue, clients, attendance per group + chart
  - Reconciliation  bookings without payment and payments without booking
  - Data            merged raw records as an Excel table

The three systems know nothing about each other. The join key is the client
e-mail + month. The "Reconciliation" sheet shows where the systems disagree,
because that is usually where cash goes missing.
"""
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.properties import PageSetupProperties

BASE = Path(__file__).parent
OUT = BASE / "output"
OUT.mkdir(exist_ok=True)

# ---------- palette ----------
BRAND = "2A78D6"      # brand blue
BRAND_D = "1C5CAB"    # one step darker
AQUA = "1BAF7A"
INK = "0B0B0B"
INK2 = "52514E"
MUTED = "898781"
TILE = "EEF4FC"       # KPI tile background
TILE_LN = "B7D3F6"    # tile border
GOOD = "0CA30C"
BAD = "D03B3B"
BAD_BG = "FDECEC"
LINE = "E1E0D9"

thin = Side(style="thin", color=TILE_LN)
TILE_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def load(name):
    with open(BASE / "input" / name, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

payments = load("payments.csv")
bookings = load("bookings.csv")
attendance = load("attendance.csv")

MONTH_NAMES = {2: "February", 3: "March", 4: "April", 5: "May", 6: "June"}
month_of = lambda iso: int(iso[5:7])

# ---------- aggregates ----------
revenue_by_month = defaultdict(float)
refunds_by_month = defaultdict(float)
for p in payments:
    m, amount = month_of(p["data"]), float(p["kwota"])
    if p["status"] == "succeeded":
        revenue_by_month[m] += amount
    elif p["status"] == "refunded":
        refunds_by_month[m] += amount

course_revenue = defaultdict(float)
course_clients = defaultdict(set)
for b in bookings:
    if b["status"] == "potwierdzona":
        course_clients[b["kurs"]].add(b["email"])

paid_bookings = set()
payments_by_email_month = defaultdict(list)
for p in payments:
    if p["status"] == "succeeded":
        payments_by_email_month[(p["email"], month_of(p["data"]))].append(p)

bookings_without_payment = []
for b in bookings:
    if b["status"] != "potwierdzona":
        continue
    key = (b["email"], month_of(b["data_startu"]))
    if payments_by_email_month.get(key):
        paid_bookings.add(b["booking_id"])
        course_revenue[b["kurs"]] += float(b["cena"])
    else:
        bookings_without_payment.append(b)

booking_email_months = {(b["email"], month_of(b["data_startu"])) for b in bookings
                        if b["status"] == "potwierdzona"}
payments_without_booking = [p for p in payments
                            if p["status"] == "succeeded"
                            and (p["email"], month_of(p["data"])) not in booking_email_months]

course_attendance = defaultdict(lambda: [0, 0])
for a in attendance:
    course_attendance[a["grupa"]][1] += 1
    if a["obecny"] == "tak":
        course_attendance[a["grupa"]][0] += 1
course_attendance_rate = {c: (present / total if total else 0)
                          for c, (present, total) in course_attendance.items()}

total_revenue = sum(revenue_by_month.values())
total_refunds = sum(refunds_by_month.values())
client_count = len({b["email"] for b in bookings if b["status"] == "potwierdzona"})
confirmed = [b for b in bookings if b["status"] == "potwierdzona"]
pct_paid = len(paid_bookings) / len(confirmed) if confirmed else 0
attendance_overall = (sum(v[0] for v in course_attendance.values())
                      / max(1, sum(v[1] for v in course_attendance.values())))

# ---------- helpers ----------
def page(ws, height=1):
    """Landscape + fit to page width (nice PDF and printout)."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = height
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

def brand_stripe(ws, to_column=13):
    """Thin brand-coloured stripe across the very top of the sheet."""
    for c in range(1, to_column + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 6

def title(ws, text, subtitle, to_column=13):
    page(ws)
    brand_stripe(ws, to_column)
    ws.cell(row=3, column=2, value=text).font = Font(name="Calibri", bold=True, size=17, color=INK)
    ws.cell(row=4, column=2, value=subtitle).font = Font(name="Calibri", size=10, color=MUTED)

def tile(ws, row, col, label, value, color=INK, number_format=None):
    """KPI tile, 3 columns x 3 rows on a merged range."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col + 2)
    for rr in range(row, row + 3):
        for cc in range(col, col + 3):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill("solid", fgColor=TILE)
            cell.border = TILE_BORDER
    label_cell = ws.cell(row=row - 1, column=col, value=label.upper())
    label_cell.font = Font(name="Calibri", size=8.5, color=MUTED, bold=True)
    value_cell = ws.cell(row=row, column=col, value=value)
    value_cell.font = Font(name="Calibri", bold=True, size=22, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        value_cell.number_format = number_format

def table(ws, ref, name, style="TableStyleMedium2", header_color=BRAND):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True,
                                      showFirstColumn=False, showLastColumn=False)
    ws.add_table(t)
    # header also painted by hand so it looks the same in Excel, Sheets and LibreOffice
    first, last = ref.split(":")
    col_from = ws[first].column
    col_to = ws[last].column
    header_row = ws[first].row
    for c in range(col_from, col_to + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10.5)
        cell.alignment = Alignment(horizontal="left")

def style_chart(chart, color=BRAND):
    chart.series[0].graphicalProperties.solidFill = color
    chart.series[0].graphicalProperties.line.noFill = True
    chart.gapWidth = 60
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=LINE))
    chart.y_axis.numFmt = "#,##0"
    chart.legend = None

def column_widths(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

wb = Workbook()

# ---------- KPI ----------
ws = wb.active
ws.title = "KPI"
ws.sheet_view.showGridLines = False
title(ws, "KPI dashboard — recurring classes",
      "Payments + bookings + attendance · February–June 2026 · fictional data (demo)")

tiles = [
    ("Net revenue", total_revenue, INK, '#,##0" zł"'),
    ("Active clients", client_count, INK, "0"),
    ("Attendance", attendance_overall, BRAND_D, "0%"),
    ("Bookings paid", pct_paid, BRAND_D, "0%"),
    ("Refunds", total_refunds, INK2, '#,##0" zł"'),
    ("Items to review", len(bookings_without_payment) + len(payments_without_booking), BAD, "0"),
]
for i, (label, value, color, fmt) in enumerate(tiles):
    row = 7 + (i // 3) * 5
    col = 2 + (i % 3) * 4
    tile(ws, row, col, label, value, color, fmt)

ws.cell(row=18, column=2,
        value="Items to review = bookings without payment + payments without a booking. Details in the Reconciliation sheet.")
ws.cell(row=18, column=2).font = Font(name="Calibri", size=9, color=MUTED, italic=True)
column_widths(ws, [2.5, 12, 12, 12, 3, 12, 12, 12, 3, 12, 12, 12])
for r in (7, 12):
    for rr in range(r, r + 3):
        ws.row_dimensions[rr].height = 18

# ---------- By month ----------
ws = wb.create_sheet("By month")
ws.sheet_view.showGridLines = False
title(ws, "Revenue by month", "Net of refunds · zł")
start = 6
ws.cell(row=start, column=2, value="Month")
ws.cell(row=start, column=3, value="Net revenue")
ws.cell(row=start, column=4, value="Refunds")
for i, m in enumerate(sorted(revenue_by_month), start=start + 1):
    ws.cell(row=i, column=2, value=MONTH_NAMES[m])
    ws.cell(row=i, column=3, value=round(revenue_by_month[m], 2)).number_format = '#,##0" zł"'
    ws.cell(row=i, column=4, value=round(refunds_by_month.get(m, 0), 2)).number_format = '#,##0" zł"'
n = len(revenue_by_month)
table(ws, f"B{start}:D{start+n}", "TabMonths")
ws.conditional_formatting.add(
    f"C{start+1}:C{start+n}",
    DataBarRule(start_type="num", start_value=0, end_type="max",
                color=BRAND, showValue=True, minLength=None, maxLength=None))
column_widths(ws, [2.5, 14, 18, 14])

chart = BarChart()
chart.title = "Net revenue by month (zł)"
chart.height, chart.width = 9, 17
data = Reference(ws, min_col=3, min_row=start, max_row=start + n)
cats = Reference(ws, min_col=2, min_row=start + 1, max_row=start + n)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
style_chart(chart)
ws.add_chart(chart, "F6")

# ---------- By course ----------
ws = wb.create_sheet("By course")
ws.sheet_view.showGridLines = False
title(ws, "Result by course", "Revenue from paid bookings · attendance from the class log")
start = 6
for c, name in enumerate(["Course", "Revenue", "Clients", "Attendance"], start=2):
    ws.cell(row=start, column=c, value=name)
courses = sorted(course_revenue, key=course_revenue.get, reverse=True)
for i, course in enumerate(courses, start=start + 1):
    ws.cell(row=i, column=2, value=course)
    ws.cell(row=i, column=3, value=round(course_revenue[course], 2)).number_format = '#,##0" zł"'
    ws.cell(row=i, column=4, value=len(course_clients[course]))
    ws.cell(row=i, column=5, value=round(course_attendance_rate.get(course, 0), 3)).number_format = "0%"
n = len(courses)
table(ws, f"B{start}:E{start+n}", "TabCourses")
ws.conditional_formatting.add(
    f"C{start+1}:C{start+n}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color=AQUA, showValue=True))
column_widths(ws, [2.5, 26, 15, 10, 12])

chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Revenue by course (zł)"
chart2.height, chart2.width = 10, 16
data = Reference(ws, min_col=3, min_row=start, max_row=start + n)
cats = Reference(ws, min_col=2, min_row=start + 1, max_row=start + n)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
style_chart(chart2, color=AQUA)
ws.add_chart(chart2, "G6")

# ---------- Reconciliation ----------
ws = wb.create_sheet("Reconciliation")
ws.sheet_view.showGridLines = False
title(ws, "System reconciliation", "Where cash usually leaks: mismatches between sign-ups and payments")

row = 6
ws.cell(row=row, column=2, value=f"Confirmed bookings without payment ({len(bookings_without_payment)})")
ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=12, color=BAD)
row += 1
for c, name in enumerate(["Booking", "Date", "Course", "Client", "E-mail", "Amount"], start=2):
    ws.cell(row=row, column=c, value=name)
start_t1 = row
for b in bookings_without_payment:
    row += 1
    ws.cell(row=row, column=2, value=b["booking_id"])
    ws.cell(row=row, column=3, value=b["data_startu"])
    ws.cell(row=row, column=4, value=b["kurs"])
    ws.cell(row=row, column=5, value=b["klient"])
    ws.cell(row=row, column=6, value=b["email"])
    ws.cell(row=row, column=7, value=float(b["cena"])).number_format = '#,##0" zł"'
table(ws, f"B{start_t1}:G{row}", "TabUnpaid", style="TableStyleLight10", header_color=BAD)

row += 3
ws.cell(row=row, column=2, value=f"Payments without a booking ({len(payments_without_booking)})")
ws.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=12, color=BRAND_D)
row += 1
for c, name in enumerate(["Payment", "Date", "Amount", "E-mail", "Description"], start=2):
    ws.cell(row=row, column=c, value=name)
start_t2 = row
for p in payments_without_booking:
    row += 1
    ws.cell(row=row, column=2, value=p["payment_id"])
    ws.cell(row=row, column=3, value=p["data"])
    ws.cell(row=row, column=4, value=float(p["kwota"])).number_format = '#,##0" zł"'
    ws.cell(row=row, column=5, value=p["email"])
    ws.cell(row=row, column=6, value=p["opis"])
table(ws, f"B{start_t2}:F{row}", "TabNoBooking", style="TableStyleLight9", header_color=BRAND_D)
column_widths(ws, [2.5, 14, 12, 24, 22, 30, 12])

# ---------- Data ----------
ws = wb.create_sheet("Data")
headers = ["Source", "ID", "Date", "E-mail", "Detail", "Amount / status"]
for c, name in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=name)
row = 1
for p in payments:
    row += 1
    for c, v in enumerate(["payments", p["payment_id"], p["data"], p["email"],
                           p["opis"], f'{p["kwota"]} ({p["status"]})'], start=1):
        ws.cell(row=row, column=c, value=v)
for b in bookings:
    row += 1
    for c, v in enumerate(["bookings", b["booking_id"], b["data_startu"], b["email"],
                           b["kurs"], f'{b["cena"]} ({b["status"]})'], start=1):
        ws.cell(row=row, column=c, value=v)
table(ws, f"A1:F{row}", "TabData")
ws.freeze_panes = "A2"
column_widths(ws, [12, 14, 12, 30, 26, 18])

wb.save(OUT / "kpi_dashboard.xlsx")
print("OK -> output/kpi_dashboard.xlsx")
print(f"  net revenue: {total_revenue:,.0f} zł | clients: {client_count} "
      f"| bookings paid: {pct_paid:.0%} | attendance: {attendance_overall:.0%}")
unpaid, unmatched = len(bookings_without_payment), len(payments_without_booking)
print(f"  reconciliation: {unpaid} booking{'' if unpaid == 1 else 's'} without payment, "
      f"{unmatched} payment{'' if unmatched == 1 else 's'} without a booking")
