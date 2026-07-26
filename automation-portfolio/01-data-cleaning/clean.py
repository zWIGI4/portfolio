#!/usr/bin/env python3
"""
Cleaning and standardizing a customer database (deliver-and-done demo).
Input:   input/clients_raw.csv        (semicolon-separated)
Output:  output/clients_clean.xlsx    (clean, sorted register)
         output/change_report.md      (what was fixed and how much)

Run:  python3 clean.py
No dependency on the client's systems — the result is a standalone file.
"""
import re
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd

BASE = Path(__file__).resolve().parent
INPUT = BASE / "input" / "clients_raw.csv"
OUTPUT_DIR = BASE / "output"
OUTPUT_XLSX = OUTPUT_DIR / "clients_clean.xlsx"
REPORT = OUTPUT_DIR / "change_report.md"

# Polish month names, as spelled out in the source export.
MONTHS = {
    "stycznia": 1, "lutego": 2, "marca": 3, "kwietnia": 4, "maja": 5, "czerwca": 6,
    "lipca": 7, "sierpnia": 8, "wrzesnia": 9, "września": 9, "pazdziernika": 10,
    "października": 10, "listopada": 11, "grudnia": 12,
}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I)


def parse_date(s):
    """Reduce assorted date formats to ISO (YYYY-MM-DD). Returns None if impossible."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    m = re.match(r"(\d{1,2})\s+([a-ząćęłńóśźż]+)\s+(\d{4})", s, re.I)
    if m and m.group(2).lower() in MONTHS:
        try:
            return datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1))).date().isoformat()
        except ValueError:
            return None
    return None


def parse_amount(s):
    """'1 234,50 zł' -> 1234.50 (float). None if empty or not a number."""
    if s is None or not str(s).strip():
        return None
    s = str(s).lower().replace("zł", "").replace("pln", "").replace("\xa0", " ").strip()
    s = s.replace(" ", "")
    if "," in s and "." in s:          # 1.234,50 -> 1234.50
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def norm_phone(s):
    """To the +48XXXXXXXXX format. None unless there are 9 national digits."""
    if not s or not str(s).strip():
        return None
    digits = re.sub(r"\D", "", str(s))
    if digits.startswith("48"):
        digits = digits[2:]
    if len(digits) == 9:
        return "+48" + digits
    return None


def norm_tax_id(s):
    """10 digits, no separators. None if not exactly 10 digits."""
    if not s or not str(s).strip():
        return None
    digits = re.sub(r"\D", "", str(s))
    return digits if len(digits) == 10 else None


def main():
    df = pd.read_csv(INPUT, sep=";", dtype=str, keep_default_na=False)
    changes = {"rows_in": len(df)}

    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        df[c] = df[c].map(lambda x: x.strip() if isinstance(x, str) else x)

    # junk rows: missing name or an invalid e-mail
    df["email_norm"] = df["email"].str.lower()
    empty = df["imie i nazwisko"].eq("") | df["email_norm"].eq("")
    bad_email = ~df["email_norm"].apply(lambda e: bool(EMAIL_RE.match(e)) if e else False)
    junk = empty | bad_email
    changes["junk_removed"] = int(junk.sum())
    df = df[~junk].copy()

    df["data zamowienia"] = df["data zamowienia"].map(parse_date)
    df["kwota"] = df["kwota"].map(parse_amount)
    df["telefon"] = df["telefon"].map(norm_phone)
    df["NIP"] = df["NIP"].map(norm_tax_id)
    df["miasto"] = df["miasto"].str.title()
    df = df.rename(columns={"email_norm": "email_clean"})
    df["email"] = df["email_clean"]
    df = df.drop(columns=["email_clean"])

    # duplicates: same e-mail + date + amount (one transaction entered twice)
    before = len(df)
    df = df.drop_duplicates(subset=["email", "data zamowienia", "kwota"], keep="first")
    changes["duplicates_removed"] = before - len(df)

    changes["missing_date"] = int(df["data zamowienia"].isna().sum())
    changes["missing_amount"] = int(df["kwota"].isna().sum())
    changes["missing_phone"] = int(df["telefon"].isna().sum())
    changes["missing_tax_id"] = int(df["NIP"].isna().sum())
    changes["rows_out"] = len(df)

    df = df.sort_values(["data zamowienia", "imie i nazwisko"], na_position="last")
    df = df.rename(columns={"imie i nazwisko": "full name", "data zamowienia": "order date",
                            "kwota": "amount", "telefon": "phone", "miasto": "city",
                            "NIP": "tax ID (NIP)"})
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Clients")
        sheet = w.sheets["Clients"]
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

        # styling: Excel table (zebra stripes + filters) and header in the brand colour
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties
        from openpyxl.worksheet.table import Table, TableStyleInfo

        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        last_col = get_column_letter(len(df.columns))
        table = Table(displayName="TabClients", ref=f"A1:{last_col}{len(df) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
        for c in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = PatternFill("solid", fgColor="2A78D6")
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10.5)
            cell.alignment = Alignment(horizontal="left")
        sheet.freeze_panes = "A2"
        for r in range(2, len(df) + 2):  # amounts as numbers with a format
            sheet.cell(row=r, column=list(df.columns).index("amount") + 1).number_format = "#,##0.00"

    total = df["amount"].dropna().sum()
    n_amounts = int(df["amount"].notna().sum())
    average = total / n_amounts if n_amounts else 0.0

    def fmt_pln(x):  # 1234.5 -> "1 234,50"
        return f"{x:,.2f}".replace(",", " ").replace(".", ",")

    REPORT.write_text(
        f"""# Data cleaning change report

Input: `{INPUT.name}` — **{changes['rows_in']} rows**
Output: `{OUTPUT_XLSX.name}` — **{changes['rows_out']} clean records**

## What was fixed
- Junk rows removed (missing name / invalid e-mail): **{changes['junk_removed']}**
- Duplicates removed (same e-mail + date + amount): **{changes['duplicates_removed']}**
- Dates normalized to ISO `YYYY-MM-DD` (from 5 notations, incl. spelled-out Polish)
- Amounts converted to numbers ("1 234,50 zł", "2200,00", "999.99" → 1234.5, 2200.0, 999.99)
- Phones unified to `+48XXXXXXXXX`
- Tax IDs to 10 digits, no separators
- E-mails lowercased, whitespace trimmed; cities capitalized

## Remaining gaps for the client to fill
- Missing date: **{changes['missing_date']}** | missing amount: **{changes['missing_amount']}** | missing phone: **{changes['missing_phone']}** | missing tax ID: **{changes['missing_tax_id']}**

## Quick KPIs from the clean data
- Order total: **{fmt_pln(total)} zł**
- Average order: **{fmt_pln(average)} zł** (from {n_amounts} records with an amount)
"""
    )
    print(f"OK: {changes['rows_in']} -> {changes['rows_out']} records")
    print(f"   removed: {changes['junk_removed']} junk rows, {changes['duplicates_removed']} duplicates")
    print(f"   files: {OUTPUT_XLSX.name}, {REPORT.name}")


if __name__ == "__main__":
    sys.exit(main())
