#!/usr/bin/env python3
"""
Extract data from PDF invoices → Excel register.
Input:  input/*.pdf  (invoices)
Output: output/invoice_register.xlsx — 2 sheets:
        • Invoices    (number, dates, seller, tax IDs, net, gross)
        • Line items  (every line from every invoice, ready to filter)

Run:  python3 extract.py
On a real job the parser is fitted to the client's invoice layout (different
issuers — different templates); this demo shows the full flow on one layout.
"""
import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
INPUT = BASE / "input"
OUTPUT = BASE / "output" / "invoice_register.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2A78D6")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def amount(s):
    """'1 234,56 zł' -> 1234.56"""
    s = s.replace("zł", "").replace("\xa0", " ").strip().replace(" ", "").replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def party_names(page):
    """Seller/buyer names from a two-column block — by coordinates.

    extract_text() glues the columns together, so we cut out the left and the
    right half of the area between the "Sprzedawca" header and the start of the
    line-item table ("Lp.") and take the first line of each half.
    """
    words = page.extract_words()
    y_top = next((w["bottom"] for w in words if w["text"] == "Sprzedawca"), None)
    y_bot = next((w["top"] for w in words if w["text"].startswith("Lp")), page.height)
    if y_top is None:
        return None, None
    middle = page.width / 2
    # within_bbox (not crop): takes only objects fully inside the area,
    # +2pt margin so the "Sprzedawca" header itself is not picked up
    left = page.within_bbox((0, y_top + 2, middle, y_bot)).extract_text() or ""
    right = page.within_bbox((middle, y_top + 2, page.width, y_bot)).extract_text() or ""
    first_line = lambda t: next((ln.strip() for ln in t.splitlines() if ln.strip()), None)
    return first_line(left), first_line(right)


def parse(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ""
        tables = page.extract_tables()
        seller_xy, buyer_xy = party_names(page)

    data = {"file": pdf_path.name}
    m = re.search(r"Faktura VAT\s+(\S+)", text)
    data["invoice_no"] = m.group(1) if m else None
    m = re.search(r"Data wystawienia:\s*([\d-]+)", text)
    data["issue_date"] = m.group(1) if m else None
    m = re.search(r"Termin płatności:\s*([\d-]+)", text)
    data["due_date"] = m.group(1) if m else None

    # seller/buyer come from the coordinate crops, not from the glued text
    tax_ids = re.findall(r"NIP\s?(\d{10})", text)
    data["seller_tax_id"] = tax_ids[0] if len(tax_ids) > 0 else None
    data["buyer_tax_id"] = tax_ids[1] if len(tax_ids) > 1 else None
    data["seller"] = seller_xy
    data["buyer"] = buyer_xy

    m = re.search(r"Do zapłaty:\s*([\d\s.,]+zł)\s*\(netto:\s*([\d\s.,]+zł)\)", text)
    data["gross"] = amount(m.group(1)) if m else None
    data["net"] = amount(m.group(2)) if m else None

    line_items = []
    for tab in tables:
        if not tab or not tab[0]:
            continue
        head_row = [c or "" for c in tab[0]]
        if "Lp." not in head_row[0]:
            continue
        for row in tab[1:]:
            if not row or not (row[0] or "").strip().isdigit():
                continue
            line_items.append({
                "invoice": data["invoice_no"],
                "no": int(row[0]),
                "item": (row[1] or "").replace("\n", " ").strip(),
                "qty": amount(row[2] or "") or (row[2] or "").strip(),
                "unit_net": amount(row[3] or ""),
                "vat": (row[4] or "").strip(),
                "gross": amount(row[5] or ""),
            })
    return data, line_items


def write_header(ws, columns):
    ws.append(columns)
    for c in ws[1]:
        c.fill, c.font = HEADER_FILL, HEADER_FONT
    ws.freeze_panes = "A2"


def as_table(ws, n_cols, n_rows, name):
    """Excel table (banded rows + filters) + number formats — looks like a product."""
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.properties import PageSetupProperties
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    for c in ws[1]:
        c.alignment = Alignment(horizontal="left")


def main():
    files = sorted(INPUT.glob("*.pdf"))
    if not files:
        print("No PDFs in input/ — run first: python3 make_sample_invoices.py")
        return
    invoices, line_items = [], []
    for p in files:
        invoice, items = parse(p)
        invoices.append(invoice)
        line_items += items
        print(f"OK {p.name}: {invoice['invoice_no']} | gross {invoice['gross']} | line items {len(items)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    invoice_keys = ["invoice_no", "issue_date", "due_date", "seller",
                    "seller_tax_id", "buyer", "buyer_tax_id", "net", "gross", "file"]
    INVOICE_LABELS = ["invoice no", "issue date", "due date", "seller", "seller tax ID",
                      "buyer", "buyer tax ID", "net", "gross", "file"]
    write_header(ws, INVOICE_LABELS)
    for invoice in invoices:
        ws.append([invoice.get(k) for k in invoice_keys])
    for i, w in enumerate([16, 15, 15, 30, 14, 22, 13, 11, 11, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    as_table(ws, len(invoice_keys), len(invoices), "TabInvoices")
    for r in range(2, len(invoices) + 2):
        for col in (8, 9):  # net, gross
            ws.cell(row=r, column=col).number_format = "#,##0.00"

    ws_items = wb.create_sheet("Line items")
    item_keys = ["invoice", "no", "item", "qty", "unit_net", "vat", "gross"]
    ITEM_LABELS = ["invoice", "no", "item", "qty", "unit net", "vat", "gross"]
    write_header(ws_items, ITEM_LABELS)
    for item in line_items:
        ws_items.append([item.get(k) for k in item_keys])
    for i, w in enumerate([16, 6, 46, 8, 12, 8, 12], 1):
        ws_items.column_dimensions[get_column_letter(i)].width = w
    as_table(ws_items, len(item_keys), len(line_items), "TabLineItems")
    for r in range(2, len(line_items) + 2):
        for col in (5, 7):  # unit_net, gross
            ws_items.cell(row=r, column=col).number_format = "#,##0.00"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    total = sum(invoice["gross"] or 0 for invoice in invoices)
    total_txt = f"{total:,.2f}".replace(",", " ")
    print(f"---\nRegister: {len(invoices)} invoices, {len(line_items)} line items, "
          f"total gross {total_txt} zł")
    print(f"File: {OUTPUT.relative_to(BASE)}")


if __name__ == "__main__":
    main()
