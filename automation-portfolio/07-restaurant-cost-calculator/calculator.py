# -*- coding: utf-8 -*-
"""Restaurant operating cost calculator (an Excel tool).

Output: output/restaurant_costs.xlsx
  - Assumptions    yellow cells to edit: rent, utilities, team, food cost,
                   average bill, opening days, guests per day
  - Costs          fixed and variable costs computed by FORMULAS from Assumptions
  - Profitability  revenue, margin, break-even (guests/day) + chart

The point of this file: the client changes numbers on Assumptions and everything
recalculates itself, including the break-even point. It is plain Excel with no
macros, so it also works in Google Sheets and LibreOffice.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.properties import PageSetupProperties

OUT = Path(__file__).parent / "output"
OUT.mkdir(exist_ok=True)

# ---------- palette ----------
BRAND = "2A78D6"
BRAND_D = "1C5CAB"
INK = "0B0B0B"
INK2 = "52514E"
MUTED = "898781"
YELLOW = "FFF2CC"
TILE = "EEF4FC"
GOOD_BG = "E7F6E7"
GOOD = "0CA30C"
LINE = "E1E0D9"
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def page(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

def brand_bar(ws, to_column=12):
    for c in range(1, to_column + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 6

def sheet_header(ws, text, subtitle):
    page(ws)
    brand_bar(ws)
    ws.cell(row=2, column=2, value="")
    ws.cell(row=3, column=2, value=text).font = Font(name="Calibri", bold=True, size=16, color=INK)
    ws.cell(row=4, column=2, value=subtitle).font = Font(name="Calibri", size=10, color=MUTED)

wb = Workbook()

# ---------------- Assumptions ----------------
assumptions = wb.active
assumptions.title = "Assumptions"
assumptions.sheet_view.showGridLines = False
sheet_header(assumptions, "Restaurant cost calculator",
             "Edit the yellow cells — costs, result and break-even recalculate on their own · sample numbers")

# (label, value, defined-name alias)
ASSUMPTION_ROWS = [
    ("Premises & utilities (monthly)", None, None),
    ("Rent", 9500, "CZYNSZ"),
    ("Utilities (electricity, gas, water)", 2800, "MEDIA"),
    ("Internet, POS, software", 450, "SOFT"),
    ("Waste, servicing, small repairs", 900, "SERWIS"),
    ("Marketing", 1200, "MARKETING"),
    ("Accounting & insurance", 850, "KSIEGOWOSC"),
    ("Team (monthly, gross with overheads)", None, None),
    ("Cooks: headcount", 3, "KUCH_N"),
    ("Cooks: cost per person", 8200, "KUCH_K"),
    ("Waiters: headcount", 4, "KEL_N"),
    ("Waiters: cost per person", 6100, "KEL_K"),
    ("Kitchen help: headcount", 2, "POM_N"),
    ("Kitchen help: cost per person", 5300, "POM_K"),
    ("Manager: cost", 9000, "MGR_K"),
    ("Sales", None, None),
    ("Average bill per guest", 68, "PARAGON"),
    ("Food cost (% of the bill)", 0.32, "FOODCOST"),
    ("Guests per day (current)", 85, "GOSCIE"),
    ("Opening days per month", 26, "DNI"),
]

row = 6
for label, value, alias in ASSUMPTION_ROWS:
    if value is None and alias is None:  # section header
        row += 1
        for cc in (2, 3):
            assumptions.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=TILE)
        cell = assumptions.cell(row=row, column=2, value=label)
        cell.font = Font(name="Calibri", bold=True, size=11, color=BRAND_D)
        row += 1
        continue
    assumptions.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=11, color=INK2)
    cell = assumptions.cell(row=row, column=3, value=value)
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.border = BORDER
    cell.font = Font(name="Calibri", bold=True, size=11, color=INK)
    cell.number_format = "0%" if alias == "FOODCOST" else "#,##0"
    wb.defined_names[alias] = DefinedName(alias, attr_text=f"'Assumptions'!$C${row}")
    row += 1

assumptions.column_dimensions["A"].width = 2.5
assumptions.column_dimensions["B"].width = 44
assumptions.column_dimensions["C"].width = 12

# ---------------- Costs ----------------
costs = wb.create_sheet("Costs")
costs.sheet_view.showGridLines = False
sheet_header(costs, "Monthly costs", "Everything computed by formulas from the Assumptions sheet")

COST_ROWS = [
    ("Fixed costs", None),
    ("Rent", "=CZYNSZ"),
    ("Utilities", "=MEDIA"),
    ("Internet & software", "=SOFT"),
    ("Servicing & repairs", "=SERWIS"),
    ("Marketing", "=MARKETING"),
    ("Accounting & insurance", "=KSIEGOWOSC"),
    ("Team", "=KUCH_N*KUCH_K+KEL_N*KEL_K+POM_N*POM_K+MGR_K"),
    ("Total fixed costs", "TOTAL_FIXED"),
    ("Variable costs", None),
    ("Ingredients (food cost)", "=GOSCIE*DNI*PARAGON*FOODCOST"),
    ("Total variable costs", "TOTAL_VARIABLE"),
    ("Total costs", "TOTAL_ALL"),
]
row = 6
fixed_range_start = None
for label, formula in COST_ROWS:
    if formula is None:
        row += 1
        for cc in (2, 3):
            costs.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=TILE)
        costs.cell(row=row, column=2, value=label).font = Font(name="Calibri", bold=True, size=11, color=BRAND_D)
        row += 1
        if label == "Fixed costs":
            fixed_range_start = row
        continue
    costs.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=11, color=INK2)
    if formula == "TOTAL_FIXED":
        cell = costs.cell(row=row, column=3, value=f"=SUM(C{fixed_range_start}:C{row-1})")
        cell.font = Font(name="Calibri", bold=True, size=11)
        costs.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=11, color=INK)
        wb.defined_names["STALE"] = DefinedName("STALE", attr_text=f"'Costs'!$C${row}")
    elif formula == "TOTAL_VARIABLE":
        cell = costs.cell(row=row, column=3, value=f"=C{row-1}")
        cell.font = Font(name="Calibri", bold=True, size=11)
        costs.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=11, color=INK)
        wb.defined_names["ZMIENNE"] = DefinedName("ZMIENNE", attr_text=f"'Costs'!$C${row}")
    elif formula == "TOTAL_ALL":
        cell = costs.cell(row=row, column=3, value="=STALE+ZMIENNE")
        cell.font = Font(name="Calibri", bold=True, size=13, color=BRAND_D)
        costs.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=12, color=BRAND_D)
        for cc in (2, 3):
            costs.cell(row=row, column=cc).fill = PatternFill("solid", fgColor=TILE)
            costs.cell(row=row, column=cc).border = BORDER
    else:
        cell = costs.cell(row=row, column=3, value=formula)
        cell.font = Font(name="Calibri", size=11)
    cell.number_format = "#,##0\" zł\""
    row += 1
costs.column_dimensions["A"].width = 2.5
costs.column_dimensions["B"].width = 34
costs.column_dimensions["C"].width = 16

# ---------------- Profitability ----------------
profit = wb.create_sheet("Profitability")
profit.sheet_view.showGridLines = False
sheet_header(profit, "Profitability and break-even",
             "Result, margin and the point where the venue breaks even")

PROFIT_ROWS = [
    ("Monthly revenue", "=GOSCIE*DNI*PARAGON", None),
    ("Total costs", "=STALE+ZMIENNE", None),
    ("Result (profit / loss)", "=C6-C7", "RESULT"),
    ("Margin per bill (after food cost)", "=PARAGON*(1-FOODCOST)", None),
    ("Break-even: guests per day", "=STALE/(PARAGON*(1-FOODCOST))/DNI", "BREAKEVEN"),
    ("Headroom vs break-even (guests/day)", "=GOSCIE-C10", None),
]
for i, (label, formula, role) in enumerate(PROFIT_ROWS, start=6):
    profit.cell(row=i, column=2, value=label).font = Font(name="Calibri", size=11, color=INK2)
    cell = profit.cell(row=i, column=3, value=formula)
    cell.number_format = "#,##0"
    cell.font = Font(name="Calibri", size=11)
    if role == "RESULT":
        cell.font = Font(name="Calibri", bold=True, size=13, color=GOOD)
        profit.cell(row=i, column=2).font = Font(name="Calibri", bold=True, size=12, color=INK)
        for cc in (2, 3):
            profit.cell(row=i, column=cc).fill = PatternFill("solid", fgColor=GOOD_BG)
            profit.cell(row=i, column=cc).border = BORDER
    elif role == "BREAKEVEN":
        cell.font = Font(name="Calibri", bold=True, size=13, color=BRAND_D)
        profit.cell(row=i, column=2).font = Font(name="Calibri", bold=True, size=12, color=INK)
        for cc in (2, 3):
            profit.cell(row=i, column=cc).fill = PatternFill("solid", fgColor=TILE)
            profit.cell(row=i, column=cc).border = BORDER

# chart data: result vs guests per day (30..150 step 10) - formulas
profit.cell(row=14, column=2, value="Guests/day").font = Font(name="Calibri", bold=True, size=10, color=MUTED)
profit.cell(row=14, column=3, value="Monthly result").font = Font(name="Calibri", bold=True, size=10, color=MUTED)
start = 15
for i, guests in enumerate(range(30, 151, 10)):
    r = start + i
    profit.cell(row=r, column=2, value=guests).font = Font(name="Calibri", size=10, color=INK2)
    cell = profit.cell(row=r, column=3, value=f"=B{r}*DNI*PARAGON*(1-FOODCOST)-STALE")
    cell.number_format = "#,##0"
    cell.font = Font(name="Calibri", size=10, color=INK2)
last_row = start + len(range(30, 151, 10)) - 1

chart = LineChart()
chart.title = "Monthly result vs guests per day"
chart.height, chart.width = 10, 17
data = Reference(profit, min_col=3, min_row=14, max_row=last_row)
cats = Reference(profit, min_col=2, min_row=15, max_row=last_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None
s = chart.series[0]
s.graphicalProperties.line = LineProperties(solidFill=BRAND, w=31750)  # ~2.5 pt
s.smooth = False
chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=LINE))
chart.y_axis.numFmt = "#,##0"
profit.add_chart(chart, "E6")

profit.column_dimensions["A"].width = 2.5
profit.column_dimensions["B"].width = 36
profit.column_dimensions["C"].width = 14

wb.save(OUT / "restaurant_costs.xlsx")
print("OK -> output/restaurant_costs.xlsx (formulas recalculate when opened in Excel)")
