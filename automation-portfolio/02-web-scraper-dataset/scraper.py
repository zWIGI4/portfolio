#!/usr/bin/env python3
"""
Web scraping -> ready-to-use dataset (deliver-and-done demo).
Source: books.toscrape.com — a public sandbox built FOR practicing scraping
(no ToS restrictions). In a real job the source and fields are swapped for the
client's own.

Output: output/products.xlsx  and  output/products.csv
Fields: title, price (float), currency, rating (1-5), availability, url

Run:  python3 scraper.py            (3 catalog pages by default)
      python3 scraper.py --pages 5
Good practices: User-Agent header, delay between requests, error handling.
"""
import argparse
import csv
import re
import time
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
OUTPUT_DIR = BASE / "output"
START = "https://books.toscrape.com/catalogue/page-1.html"
UA = {"User-Agent": "Mozilla/5.0 (compatible; portfolio-demo-scraper/1.0)"}
RATINGS = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}


def scrape(pages):
    url = START
    rows = []
    for i in range(pages):
        r = requests.get(url, headers=UA, timeout=20)
        r.raise_for_status()
        r.encoding = "utf-8"  # the page is UTF-8; without this £ is read as "Â£"
        soup = BeautifulSoup(r.text, "html.parser")
        for art in soup.select("article.product_pod"):
            title = art.h3.a["title"].strip()
            price_txt = art.select_one("p.price_color").get_text(strip=True)  # e.g. "£51.77"
            m = re.search(r"\d+[.,]?\d*", price_txt)
            price = float(m.group().replace(",", ".")) if m else None
            currency = next((sign for sign in "£€$" if sign in price_txt), "")
            classes = art.select_one("p.star-rating")["class"]
            rating = next((RATINGS[k] for k in classes if k in RATINGS), None)
            availability = art.select_one("p.instock.availability").get_text(strip=True)
            link = urljoin(url, art.h3.a["href"])
            rows.append({
                "title": title, "price": price, "currency": currency,
                "rating_1_5": rating, "availability": availability, "url": link,
            })
        next_page = soup.select_one("li.next a")
        if not next_page:
            break
        url = urljoin(url, next_page["href"])
        time.sleep(0.3)  # polite delay between pages
    return rows


def save(rows):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fields = ["title", "price", "currency", "rating_1_5", "availability", "url"]

    with open(OUTPUT_DIR / "products.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(fields)
    for r in rows:
        ws.append([r[p] for p in fields])

    # styling: Excel table (zebra stripes + filters), header in the brand colour,
    # prices with a number format
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    last_col = get_column_letter(len(fields))
    table = Table(displayName="TabProducts", ref=f"A1:{last_col}{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="2A78D6")
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10.5)
        c.alignment = Alignment(horizontal="left")
    price_col = fields.index("price") + 1
    for r in range(2, len(rows) + 2):
        ws.cell(row=r, column=price_col).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    for i, p in enumerate(fields, 1):
        width = max([len(p)] + [len(str(r[p])) for r in rows]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 60)
    wb.save(OUTPUT_DIR / "products.xlsx")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pages", type=int, default=3)
    a = ap.parse_args()
    rows = scrape(a.pages)
    save(rows)
    prices = [r["price"] for r in rows]
    print(f"OK: collected {len(rows)} products from {a.pages} pages")
    if prices:
        print(f"   price min/avg/max: {min(prices):.2f} / {sum(prices)/len(prices):.2f} / {max(prices):.2f}")
    print(f"   files: output/products.xlsx, output/products.csv")


if __name__ == "__main__":
    main()
