#!/usr/bin/env python3
"""
Dashboard sprzedaży z surowego CSV (demo „dostarcz-i-koniec").
Wejście: dane_wejsciowe/sprzedaz.csv (surowe transakcje)
Wyjście: wyniki/dashboard.xlsx — arkusze:
         • KPI (kafelki: przychód, liczba zamówień, średnia wartość, top kategoria)
         • Wg miesiąca (tabela + wykres słupkowy)
         • Wg kategorii (tabela + wykres)
         • Dane (surowe, z filtrem)

Uruchomienie:  python3 dashboard.py
Wynik to jeden samodzielny plik .xlsx — u klienta nic nie instalujemy.
"""
from pathlib import Path
from collections import defaultdict
import csv

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BAZA = Path(__file__).resolve().parent
WEJSCIE = BAZA / "dane_wejsciowe" / "sprzedaz.csv"
WYJSCIE = BAZA / "wyniki" / "dashboard.xlsx"

MIES = ["", "Sty", "Lut", "Mar", "Kwi", "Maj", "Cze", "Lip", "Sie", "Wrz", "Paź", "Lis", "Gru"]
NAGL = PatternFill("solid", fgColor="1F3864")
NAGL_F = Font(bold=True, color="FFFFFF")
KAFEL = PatternFill("solid", fgColor="E8EEF7")
CIENKA = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def wczytaj():
    with open(WEJSCIE, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def naglowek(ws, wiersz, kolumny):
    for j, tytul in enumerate(kolumny, 1):
        c = ws.cell(wiersz, j, tytul)
        c.fill, c.font, c.border = NAGL, NAGL_F, CIENKA


def main():
    dane = wczytaj()
    for r in dane:
        r["wartosc"] = float(r["wartosc"])
        r["ilosc"] = int(r["ilosc"])
        r["miesiac"] = int(r["data"][5:7])

    przychod = sum(r["wartosc"] for r in dane)
    zamowien = len(dane)
    srednia = przychod / zamowien if zamowien else 0
    wg_kat = defaultdict(float)
    wg_mies = defaultdict(float)
    for r in dane:
        wg_kat[r["kategoria"]] += r["wartosc"]
        wg_mies[r["miesiac"]] += r["wartosc"]
    top_kat = max(wg_kat, key=wg_kat.get)

    wb = Workbook()

    # --- KPI ---
    kpi = wb.active
    kpi.title = "KPI"
    kpi["B2"] = "Dashboard sprzedaży 2025"
    kpi["B2"].font = Font(bold=True, size=16, color="1F3864")
    kafelki = [
        ("Przychód łącznie", f"{przychod:,.0f} zł".replace(",", " ")),
        ("Liczba zamówień", f"{zamowien}"),
        ("Średnia wartość zamówienia", f"{srednia:,.0f} zł".replace(",", " ")),
        ("Top kategoria", f"{top_kat}"),
    ]
    for i, (ety, war) in enumerate(kafelki):
        kol = 2 + i * 2
        e = kpi.cell(4, kol, ety); e.font = Font(size=10, color="555555")
        w = kpi.cell(5, kol, war); w.font = Font(bold=True, size=18, color="1F3864")
        for rr in (4, 5):
            for cc in (kol, kol + 1):
                kpi.cell(rr, cc).fill = KAFEL
        kpi.merge_cells(start_row=4, start_column=kol, end_row=4, end_column=kol + 1)
        kpi.merge_cells(start_row=5, start_column=kol, end_row=5, end_column=kol + 1)
        kpi.column_dimensions[get_column_letter(kol)].width = 16
        kpi.column_dimensions[get_column_letter(kol + 1)].width = 8

    # --- Wg miesiąca ---
    wm = wb.create_sheet("Wg miesiąca")
    naglowek(wm, 1, ["Miesiąc", "Przychód"])
    for m in range(1, 13):
        wm.cell(m + 1, 1, MIES[m])
        wm.cell(m + 1, 2, round(wg_mies.get(m, 0), 2)).number_format = "# ##0"
    wm.column_dimensions["A"].width = 12
    wm.column_dimensions["B"].width = 14
    wykres = BarChart()
    wykres.title = "Przychód wg miesiąca"
    wykres.type = "col"
    wykres.legend = None
    dane_ref = Reference(wm, min_col=2, min_row=1, max_row=13)
    kat_ref = Reference(wm, min_col=1, min_row=2, max_row=13)
    wykres.add_data(dane_ref, titles_from_data=True)
    wykres.set_categories(kat_ref)
    wykres.height, wykres.width = 8, 18
    wm.add_chart(wykres, "D2")

    # --- Wg kategorii ---
    wk = wb.create_sheet("Wg kategorii")
    naglowek(wk, 1, ["Kategoria", "Przychód", "Udział %"])
    for i, (k, v) in enumerate(sorted(wg_kat.items(), key=lambda x: -x[1]), start=2):
        wk.cell(i, 1, k)
        wk.cell(i, 2, round(v, 2)).number_format = "# ##0"
        wk.cell(i, 3, v / przychod).number_format = "0,0%"
    wk.column_dimensions["A"].width = 16
    wk.column_dimensions["B"].width = 14
    wk.column_dimensions["C"].width = 10
    wyk2 = BarChart()
    wyk2.title = "Przychód wg kategorii"
    wyk2.type = "bar"
    wyk2.legend = None
    d2 = Reference(wk, min_col=2, min_row=1, max_row=1 + len(wg_kat))
    k2 = Reference(wk, min_col=1, min_row=2, max_row=1 + len(wg_kat))
    wyk2.add_data(d2, titles_from_data=True)
    wyk2.set_categories(k2)
    wyk2.height, wyk2.width = 8, 18
    wk.add_chart(wyk2, "E2")

    # --- Dane surowe ---
    ds = wb.create_sheet("Dane")
    kolumny = ["nr_zamowienia", "data", "kategoria", "kanal", "ilosc", "cena_jedn", "wartosc"]
    naglowek(ds, 1, kolumny)
    for r in dane:
        ds.append([r[c] for c in kolumny])
    ds.freeze_panes = "A2"
    ds.auto_filter.ref = f"A1:G{len(dane) + 1}"
    szer = [16, 12, 14, 14, 8, 11, 11]
    for i, w in enumerate(szer, 1):
        ds.column_dimensions[get_column_letter(i)].width = w

    WYJSCIE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WYJSCIE)
    print(f"OK: dashboard z {zamowien} transakcji")
    print(f"   przychód {przychod:,.0f} zł | top kategoria: {top_kat}".replace(",", " "))
    print(f"   plik: wyniki/dashboard.xlsx (4 arkusze, 2 wykresy)")


if __name__ == "__main__":
    main()
