#!/usr/bin/env python3
"""
Web scraping -> gotowy zbiór danych (demo „dostarcz-i-koniec").
Źródło: books.toscrape.com — publiczny sandbox stworzony DO ćwiczenia scrapingu
(brak ograniczeń ToS). W realnym zleceniu podmieniam źródło i pola pod klienta.

Wyjście: wyniki/produkty.xlsx  oraz  wyniki/produkty.csv
Pola: tytuł, cena (float), waluta, ocena (1-5), dostępność, url

Uruchomienie:  python3 scraper.py            (domyślnie 3 strony katalogu)
               python3 scraper.py --strony 5
Dobre praktyki: nagłówek User-Agent, opóźnienie między żądaniami, obsługa błędów.
"""
import argparse
import csv
import re
import time
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BAZA = Path(__file__).resolve().parent
START = "https://books.toscrape.com/catalogue/page-1.html"
UA = {"User-Agent": "Mozilla/5.0 (kompatybilny; portfolio-demo-scraper/1.0)"}
OCENY = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}


def scrapuj(strony):
    url = START
    wiersze = []
    for i in range(strony):
        r = requests.get(url, headers=UA, timeout=20)
        r.raise_for_status()
        r.encoding = "utf-8"  # strona jest w UTF-8; bez tego £ czyta się jako „Â£"
        zupa = BeautifulSoup(r.text, "html.parser")
        for art in zupa.select("article.product_pod"):
            tytul = art.h3.a["title"].strip()
            cena_txt = art.select_one("p.price_color").get_text(strip=True)  # np. "£51.77"
            m = re.search(r"\d+[.,]?\d*", cena_txt)
            cena = float(m.group().replace(",", ".")) if m else None
            waluta = next((z for z in "£€$" if z in cena_txt), "")
            klasy = art.select_one("p.star-rating")["class"]
            ocena = next((OCENY[k] for k in klasy if k in OCENY), None)
            dostepnosc = art.select_one("p.instock.availability").get_text(strip=True)
            link = urljoin(url, art.h3.a["href"])
            wiersze.append({
                "tytul": tytul, "cena": cena, "waluta": waluta,
                "ocena_1_5": ocena, "dostepnosc": dostepnosc, "url": link,
            })
        nast = zupa.select_one("li.next a")
        if not nast:
            break
        url = urljoin(url, nast["href"])
        time.sleep(0.3)  # uprzejme opóźnienie między stronami
    return wiersze


def zapisz(wiersze):
    (BAZA / "wyniki").mkdir(parents=True, exist_ok=True)
    pola = ["tytul", "cena", "waluta", "ocena_1_5", "dostepnosc", "url"]

    with open(BAZA / "wyniki" / "produkty.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=pola)
        w.writeheader()
        w.writerows(wiersze)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produkty"
    ws.append(pola)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in wiersze:
        ws.append([r[p] for p in pola])
    ws.freeze_panes = "A2"
    for i, p in enumerate(pola, 1):
        dl = max([len(p)] + [len(str(r[p])) for r in wiersze]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(dl, 60)
    wb.save(BAZA / "wyniki" / "produkty.xlsx")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--strony", type=int, default=3)
    a = ap.parse_args()
    wiersze = scrapuj(a.strony)
    zapisz(wiersze)
    ceny = [r["cena"] for r in wiersze]
    print(f"OK: zebrano {len(wiersze)} produktów z {a.strony} stron")
    if ceny:
        print(f"   cena min/śr/max: {min(ceny):.2f} / {sum(ceny)/len(ceny):.2f} / {max(ceny):.2f}")
    print(f"   pliki: wyniki/produkty.xlsx, wyniki/produkty.csv")


if __name__ == "__main__":
    main()
