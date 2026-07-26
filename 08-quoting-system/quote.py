# -*- coding: utf-8 -*-
"""Quoting system — product catalogue + a request turn into a live Excel quote.

Input:  input/catalogue.csv      products, prices, unit costs
        input/quote_request.csv  what the customer asked for
Output: output/quotation.xlsx
  - Quote            the working sheet: product dropdowns, quantities and
                     discounts in yellow cells, everything else by FORMULAS
  - Catalogue        price list feeding the dropdowns and the lookups
  - Term comparison  12 / 24 / 36 / 48 months side by side + chart
  - Summary          contract value, cost, gross margin + chart
  - Customer quote   print-ready A4 page, no cost or margin anywhere

The point of this file: the sales rep edits quantities, discounts or the
contract term and every total, the margin and the customer-facing page
recalculate inside Excel. Plain formulas, no macros, so it also works in
LibreOffice and Google Sheets.
"""
import csv
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

HERE = Path(__file__).parent
IN = HERE / "input"
OUT = HERE / "output"
OUT.mkdir(exist_ok=True)

# ---------- palette ----------
BRAND = "2A78D6"
BRAND_D = "1C5CAB"
INK = "0B0B0B"
INK2 = "52514E"
MUTED = "898781"
YELLOW = "FFF2CC"
TILE = "EEF4FC"
GOOD_BG = "E7F6E7"
GOOD = "0CA30C"
LINE = "E1E0D9"
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '#,##0.00" zł"'
PCT = "0.0%"
DATE_FMT = "dd.mm.yyyy"

# quote settings (the yellow cells the rep edits)
TERM_MONTHS = 24
VAT_RATE = 0.23
CUSTOMER = "Kowalski Logistyka Sp. z o.o."
CUSTOMER_CITY = "Poznań"
CUSTOMER_CONTACT = "Anna Nowak"
QUOTE_NO = "OFE-2026-041"
SELLER = "Telko Partner Sp. z o.o."
VALID_DAYS = 30


def page(ws, orientation="landscape"):
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def brand_bar(ws, to_column=13):
    for c in range(1, to_column + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 6


def sheet_header(ws, text, subtitle, to_column=13, orientation="landscape"):
    page(ws, orientation)
    brand_bar(ws, to_column)
    ws.cell(row=3, column=2, value=text).font = Font(name="Calibri", bold=True, size=16, color=INK)
    ws.cell(row=4, column=2, value=subtitle).font = Font(name="Calibri", size=10, color=MUTED)


def label(ws, row, col, text, bold=False, color=INK2, size=11):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", bold=bold, size=size, color=color)
    return cell


def input_cell(ws, row, col, value, fmt=None):
    """A yellow, unlocked cell — this is what the user is meant to edit."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.border = BORDER
    cell.protection = Protection(locked=False)
    cell.font = Font(name="Calibri", bold=True, size=11, color=INK)
    if fmt:
        cell.number_format = fmt
    return cell


def table_header(ws, row, headers, start_col=2):
    for i, head in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=head)
        cell.fill = PatternFill("solid", fgColor=BRAND_D)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


# ---------------- read the inputs ----------------
with open(IN / "catalogue.csv", encoding="utf-8") as f:
    catalogue = list(csv.DictReader(f))
with open(IN / "quote_request.csv", encoding="utf-8") as f:
    request = list(csv.DictReader(f))

price_of = {r["product"]: float(r["unit_price"]) for r in catalogue}
cost_of = {r["product"]: float(r["unit_cost"]) for r in catalogue}
billing_of = {r["product"]: r["billing"] for r in catalogue}

wb = Workbook()

# ---------------- Catalogue ----------------
cat = wb.active
cat.title = "Catalogue"
cat.sheet_view.showGridLines = False
sheet_header(cat, "Product catalogue", "Feeds the dropdowns and every lookup on the Quote sheet", to_column=7)

CAT_HEAD_ROW = 6
CAT_FIRST = CAT_HEAD_ROW + 1
table_header(cat, CAT_HEAD_ROW, ["SKU", "Product", "Category", "Billing", "Unit price", "Unit cost"])

for i, item in enumerate(catalogue):
    r = CAT_FIRST + i
    label(cat, r, 2, item["sku"], size=10)
    label(cat, r, 3, item["product"], size=10, color=INK)
    label(cat, r, 4, item["category"], size=10)
    label(cat, r, 5, item["billing"], size=10)
    for col, key in ((6, "unit_price"), (7, "unit_cost")):
        cell = cat.cell(row=r, column=col, value=float(item[key]))
        cell.number_format = MONEY
        cell.font = Font(name="Calibri", size=10, color=INK2)
    for col in range(2, 8):
        cat.cell(row=r, column=col).border = BORDER

CAT_LAST = CAT_FIRST + len(catalogue) - 1
for name, col in (("CAT_SKU", "B"), ("CAT_PRODUCT", "C"), ("CAT_CATEGORY", "D"),
                  ("CAT_BILLING", "E"), ("CAT_PRICE", "F"), ("CAT_COST", "G")):
    wb.defined_names[name] = DefinedName(
        name, attr_text=f"'Catalogue'!${col}${CAT_FIRST}:${col}${CAT_LAST}")

for col, width in (("A", 2.5), ("B", 16), ("C", 36), ("D", 14), ("E", 11), ("F", 13), ("G", 13)):
    cat.column_dimensions[col].width = width
cat.freeze_panes = f"A{CAT_FIRST}"

# ---------------- Quote ----------------
q = wb.create_sheet("Quote")
q.sheet_view.showGridLines = False
sheet_header(q, "Quotation builder",
             "Edit the yellow cells — totals, margin and the customer page recalculate on their own")

today = date.today()
label(q, 6, 2, "Customer", bold=True, color=BRAND_D)
label(q, 7, 2, "City")
label(q, 8, 2, "Contact person")
label(q, 9, 2, "Contract term (months)")
label(q, 10, 2, "VAT rate")
input_cell(q, 6, 3, CUSTOMER)
input_cell(q, 7, 3, CUSTOMER_CITY)
input_cell(q, 8, 3, CUSTOMER_CONTACT)
input_cell(q, 9, 3, TERM_MONTHS, "0")
input_cell(q, 10, 3, VAT_RATE, "0%")
wb.defined_names["TERM"] = DefinedName("TERM", attr_text="'Quote'!$C$9")
wb.defined_names["VAT_RATE"] = DefinedName("VAT_RATE", attr_text="'Quote'!$C$10")

label(q, 6, 6, "Quote no.", bold=True, color=BRAND_D)
label(q, 7, 6, "Date")
label(q, 8, 6, "Valid until")
label(q, 9, 6, "Seller")
label(q, 6, 8, QUOTE_NO, bold=True, color=INK)
c = q.cell(row=7, column=8, value=today)
c.number_format = DATE_FMT
c.font = Font(name="Calibri", size=11, color=INK2)
c = q.cell(row=8, column=8, value=today + timedelta(days=VALID_DAYS))
c.number_format = DATE_FMT
c.font = Font(name="Calibri", size=11, color=INK2)
label(q, 9, 8, SELLER)

HEAD_ROW = 13
FIRST = HEAD_ROW + 1
LAST = FIRST + len(request) - 1
table_header(q, HEAD_ROW, [
    "Product", "SKU", "Category", "Billing", "Qty", "Unit price",
    "Disc. %", "Net one-off", "Net monthly", "Cost (contract)", "Margin %",
], start_col=3)

for i, line in enumerate(request):
    r = FIRST + i
    input_cell(q, r, 3, line["product"]).font = Font(name="Calibri", size=10, color=INK)
    lookup = f'MATCH($C{r},CAT_PRODUCT,0)'
    q.cell(row=r, column=4, value=f'=IFERROR(INDEX(CAT_SKU,{lookup}),"")')
    q.cell(row=r, column=5, value=f'=IFERROR(INDEX(CAT_CATEGORY,{lookup}),"")')
    q.cell(row=r, column=6, value=f'=IFERROR(INDEX(CAT_BILLING,{lookup}),"")')
    input_cell(q, r, 7, int(line["quantity"]), "0").font = Font(name="Calibri", size=10, color=INK)
    q.cell(row=r, column=8, value=f'=IFERROR(INDEX(CAT_PRICE,{lookup}),0)').number_format = MONEY
    input_cell(q, r, 9, float(line["discount"]), "0%").font = Font(name="Calibri", size=10, color=INK)
    q.cell(row=r, column=10,
           value=f'=IF($F{r}="one-off",ROUND($G{r}*$H{r}*(1-$I{r}),2),0)').number_format = MONEY
    q.cell(row=r, column=11,
           value=f'=IF($F{r}="monthly",ROUND($G{r}*$H{r}*(1-$I{r}),2),0)').number_format = MONEY
    q.cell(row=r, column=12,
           value=f'=IFERROR(IF($F{r}="monthly",$G{r}*INDEX(CAT_COST,{lookup})*TERM,'
                 f'$G{r}*INDEX(CAT_COST,{lookup})),0)').number_format = MONEY
    q.cell(row=r, column=13,
           value=f'=IFERROR(($J{r}+$K{r}*TERM-$L{r})/($J{r}+$K{r}*TERM),"")').number_format = PCT
    for col in range(3, 14):
        cell = q.cell(row=r, column=col)
        cell.border = BORDER
        if cell.font.size is None or col not in (3, 7, 9):
            cell.font = Font(name="Calibri", size=10, color=INK2)

# product dropdown, driven by the catalogue
dv = DataValidation(type="list", formula1=f"'Catalogue'!$C${CAT_FIRST}:$C${CAT_LAST}",
                    allow_blank=True, showDropDown=False)
dv.error = "Pick a product from the catalogue."
dv.errorTitle = "Unknown product"
q.add_data_validation(dv)
dv.add(f"C{FIRST}:C{LAST}")

TOT = LAST + 1
label(q, TOT, 2, "Total", bold=True, color=INK, size=11)
for col, letter in ((10, "J"), (11, "K"), (12, "L")):
    cell = q.cell(row=TOT, column=col, value=f"=SUM({letter}{FIRST}:{letter}{LAST})")
    cell.number_format = MONEY
    cell.font = Font(name="Calibri", bold=True, size=11, color=INK)
for col in range(2, 14):
    q.cell(row=TOT, column=col).fill = PatternFill("solid", fgColor=TILE)
    q.cell(row=TOT, column=col).border = BORDER

wb.defined_names["TOTAL_ONEOFF"] = DefinedName("TOTAL_ONEOFF", attr_text=f"'Quote'!$J${TOT}")
wb.defined_names["TOTAL_MONTHLY"] = DefinedName("TOTAL_MONTHLY", attr_text=f"'Quote'!$K${TOT}")
wb.defined_names["TOTAL_COST"] = DefinedName("TOTAL_COST", attr_text=f"'Quote'!$L${TOT}")

SUM_ROW = TOT + 2
SUMMARY_ROWS = [
    ("Upfront, net (one-off items)", "=TOTAL_ONEOFF", MONEY, "UPFRONT_NET"),
    ("Recurring, net (per month)", "=TOTAL_MONTHLY", MONEY, "MONTHLY_NET"),
    ("Contract value, net", "=TOTAL_ONEOFF+TOTAL_MONTHLY*TERM", MONEY, "CONTRACT_NET"),
    ("VAT", "=ROUND(CONTRACT_NET*VAT_RATE,2)", MONEY, "CONTRACT_VAT"),
    ("Contract value, gross", "=CONTRACT_NET+CONTRACT_VAT", MONEY, "CONTRACT_GROSS"),
    ("Cost of delivery", "=TOTAL_COST", MONEY, None),
    ("Gross margin", "=CONTRACT_NET-TOTAL_COST", MONEY, "MARGIN_VALUE"),
    ("Gross margin %", "=IFERROR(MARGIN_VALUE/CONTRACT_NET,0)", PCT, "MARGIN_PCT"),
]
label(q, SUM_ROW - 1, 2, "Quote summary", bold=True, color=BRAND_D)
for i, (text, formula, fmt, name) in enumerate(SUMMARY_ROWS):
    r = SUM_ROW + i
    label(q, r, 2, text)
    cell = q.cell(row=r, column=4, value=formula)
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=True, size=11, color=INK)
    cell.border = BORDER
    if name:
        wb.defined_names[name] = DefinedName(name, attr_text=f"'Quote'!$D${r}")
    if text == "Contract value, gross":
        for cc in (2, 3, 4):
            q.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TILE)
        cell.font = Font(name="Calibri", bold=True, size=13, color=BRAND_D)
    if text == "Gross margin %":
        for cc in (2, 3, 4):
            q.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=GOOD_BG)
        cell.font = Font(name="Calibri", bold=True, size=13, color=GOOD)

for col, width in (("A", 2.5), ("B", 30), ("C", 34), ("D", 15), ("E", 13), ("F", 11),
                   ("G", 7), ("H", 12), ("I", 9), ("J", 13), ("K", 13), ("L", 14), ("M", 10)):
    q.column_dimensions[col].width = width
q.freeze_panes = f"A{FIRST}"
q.protection.sheet = True
q.protection.enable()

# ---------------- Term comparison ----------------
cmp_ws = wb.create_sheet("Term comparison")
cmp_ws.sheet_view.showGridLines = False
sheet_header(cmp_ws, "Contract term comparison",
             "The same basket priced over four terms — the selected one is marked", to_column=8)

TERMS = [12, 24, 36, 48]
CMP_HEAD = 6
table_header(cmp_ws, CMP_HEAD, ["Term (months)", "Upfront, net", "Recurring total, net",
                                "Contract value, net", "VAT", "Contract value, gross",
                                "Monthly if financed", ""])
for i, term in enumerate(TERMS):
    r = CMP_HEAD + 1 + i
    cell = cmp_ws.cell(row=r, column=2, value=term)
    cell.font = Font(name="Calibri", bold=True, size=11, color=INK)
    cell.alignment = Alignment(horizontal="center")
    cmp_ws.cell(row=r, column=3, value="=TOTAL_ONEOFF")
    cmp_ws.cell(row=r, column=4, value=f"=TOTAL_MONTHLY*$B{r}")
    cmp_ws.cell(row=r, column=5, value=f"=TOTAL_ONEOFF+TOTAL_MONTHLY*$B{r}")
    cmp_ws.cell(row=r, column=6, value=f"=ROUND($E{r}*VAT_RATE,2)")
    cmp_ws.cell(row=r, column=7, value=f"=$E{r}+$F{r}")
    cmp_ws.cell(row=r, column=8, value=f"=$E{r}/$B{r}")
    marker = cmp_ws.cell(row=r, column=9, value=f'=IF($B{r}=TERM,"< selected term","")')
    marker.font = Font(name="Calibri", bold=True, size=10, color=BRAND_D)
    for col in range(3, 9):
        cell = cmp_ws.cell(row=r, column=col)
        cell.number_format = MONEY
        cell.font = Font(name="Calibri", size=11, color=INK2)
        cell.border = BORDER
    cmp_ws.cell(row=r, column=2).border = BORDER

CMP_LAST = CMP_HEAD + len(TERMS)
chart = BarChart()
chart.type = "col"
chart.title = "Contract value, net — by term"
chart.height, chart.width = 9, 16
data = Reference(cmp_ws, min_col=5, min_row=CMP_HEAD, max_row=CMP_LAST)
cats = Reference(cmp_ws, min_col=2, min_row=CMP_HEAD + 1, max_row=CMP_LAST)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.legend = None
chart.y_axis.numFmt = "#,##0"
chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=LINE))
chart.series[0].graphicalProperties.solidFill = BRAND
cmp_ws.add_chart(chart, "B13")

for col, width in (("A", 2.5), ("B", 15), ("C", 14), ("D", 19), ("E", 18),
                   ("F", 13), ("G", 19), ("H", 18), ("I", 17)):
    cmp_ws.column_dimensions[col].width = width

# ---------------- Summary ----------------
s = wb.create_sheet("Summary")
s.sheet_view.showGridLines = False
sheet_header(s, "Quote summary", "Everything below is live — it follows the Quote sheet", to_column=8)

TILES = [
    ("Contract value, net", "=CONTRACT_NET", MONEY),
    ("VAT", "=CONTRACT_VAT", MONEY),
    ("Contract value, gross", "=CONTRACT_GROSS", MONEY),
    ("Upfront, net", "=UPFRONT_NET", MONEY),
    ("Recurring, net / month", "=MONTHLY_NET", MONEY),
    ("Contract term (months)", "=TERM", "0"),
    ("Cost of delivery", "=TOTAL_COST", MONEY),
    ("Gross margin", "=MARGIN_VALUE", MONEY),
    ("Gross margin %", "=MARGIN_PCT", PCT),
]
row = 6
for i, (text, formula, fmt) in enumerate(TILES):
    col = 2 + (i % 3) * 2
    if i and i % 3 == 0:
        row += 3
    s.cell(row=row, column=col, value=text).font = Font(name="Calibri", bold=True, size=9, color=MUTED)
    cell = s.cell(row=row + 1, column=col, value=formula)
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=True, size=15, color=BRAND_D)
    for rr in (row, row + 1):
        for cc in (col, col + 1):
            s.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=TILE)
            s.cell(row=rr, column=cc).border = BORDER

CHART_ROW = row + 4
s.cell(row=CHART_ROW, column=2, value="Metric").font = Font(name="Calibri", bold=True, size=10, color=MUTED)
s.cell(row=CHART_ROW, column=3, value="Value").font = Font(name="Calibri", bold=True, size=10, color=MUTED)
for i, (text, formula) in enumerate([("Revenue (net)", "=CONTRACT_NET"),
                                     ("Cost", "=TOTAL_COST"),
                                     ("Gross margin", "=MARGIN_VALUE")]):
    r = CHART_ROW + 1 + i
    s.cell(row=r, column=2, value=text).font = Font(name="Calibri", size=10, color=INK2)
    cell = s.cell(row=r, column=3, value=formula)
    cell.number_format = MONEY
    cell.font = Font(name="Calibri", size=10, color=INK2)

chart2 = BarChart()
chart2.type = "col"
chart2.title = "Revenue / cost / margin over the contract"
chart2.height, chart2.width = 8, 15
data = Reference(s, min_col=3, min_row=CHART_ROW, max_row=CHART_ROW + 3)
cats = Reference(s, min_col=2, min_row=CHART_ROW + 1, max_row=CHART_ROW + 3)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.legend = None
chart2.y_axis.numFmt = "#,##0"
chart2.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=LINE))
chart2.series[0].graphicalProperties.solidFill = BRAND
s.add_chart(chart2, "F" + str(CHART_ROW))

for col, width in (("A", 2.5), ("B", 22), ("C", 16), ("D", 22), ("E", 16), ("F", 22), ("G", 16)):
    s.column_dimensions[col].width = width

# ---------------- Customer quote ----------------
cq = wb.create_sheet("Customer quote")
cq.sheet_view.showGridLines = False
sheet_header(cq, "Quotation", "Print-ready page for the customer — no cost or margin data",
             to_column=7, orientation="portrait")

label(cq, 6, 2, SELLER, bold=True, color=INK, size=12)
label(cq, 7, 2, "ul. Przemysłowa 14, 61-441 Poznań")
label(cq, 8, 2, "NIP 7792345678")
label(cq, 6, 6, "Quote no.", bold=True, color=BRAND_D)
label(cq, 7, 6, "Date")
label(cq, 8, 6, "Valid until")
cq.cell(row=6, column=7, value="='Quote'!$H$6").font = Font(name="Calibri", bold=True, size=11, color=BRAND_D)
for src_row, dst_row in ((7, 7), (8, 8)):
    c = cq.cell(row=dst_row, column=7, value=f"='Quote'!$H${src_row}")
    c.number_format = DATE_FMT
    c.font = Font(name="Calibri", size=10, color=INK2)

label(cq, 10, 2, "Prepared for", bold=True, color=BRAND_D)
cq.cell(row=11, column=2, value="='Quote'!$C$6").font = Font(name="Calibri", bold=True, size=11, color=INK)
cq.cell(row=12, column=2, value="='Quote'!$C$7").font = Font(name="Calibri", size=10, color=INK2)
cq.cell(row=13, column=2, value="=\"Attn: \"&'Quote'!$C$8").font = Font(name="Calibri", size=10, color=INK2)

CQ_HEAD = 15
table_header(cq, CQ_HEAD, ["#", "Product", "Billing", "Qty", "Unit price", "Net"])
for i in range(len(request)):
    r = CQ_HEAD + 1 + i
    src = FIRST + i
    label(cq, r, 2, i + 1, size=10, color=MUTED)
    cq.cell(row=r, column=3, value=f"='Quote'!$C${src}").font = Font(name="Calibri", size=10, color=INK)
    cq.cell(row=r, column=4, value=f"='Quote'!$F${src}").font = Font(name="Calibri", size=10, color=INK2)
    cq.cell(row=r, column=5, value=f"='Quote'!$G${src}").font = Font(name="Calibri", size=10, color=INK2)
    cell = cq.cell(row=r, column=6, value=f"='Quote'!$H${src}*(1-'Quote'!$I${src})")
    cell.number_format = MONEY
    cell.font = Font(name="Calibri", size=10, color=INK2)
    cell = cq.cell(row=r, column=7, value=f"='Quote'!$J${src}+'Quote'!$K${src}")
    cell.number_format = MONEY
    cell.font = Font(name="Calibri", size=10, color=INK2)
    for col in range(2, 8):
        cq.cell(row=r, column=col).border = BORDER

CQ_TOT = CQ_HEAD + len(request) + 2
CQ_ROWS = [
    ("Upfront, net", "=UPFRONT_NET", MONEY),
    ("Recurring, net / month", "=MONTHLY_NET", MONEY),
    ("Contract term (months)", "=TERM", "0"),
    ("Contract value, net", "=CONTRACT_NET", MONEY),
    ("VAT 23%", "=CONTRACT_VAT", MONEY),
    ("Contract value, gross", "=CONTRACT_GROSS", MONEY),
]
for i, (text, formula, fmt) in enumerate(CQ_ROWS):
    r = CQ_TOT + i
    label(cq, r, 5, text, bold=(i == len(CQ_ROWS) - 1))
    cell = cq.cell(row=r, column=7, value=formula)
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=True, size=11, color=INK)
    cell.border = BORDER
    if i == len(CQ_ROWS) - 1:
        for cc in (5, 6, 7):
            cq.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=TILE)
        cell.font = Font(name="Calibri", bold=True, size=13, color=BRAND_D)

FOOT = CQ_TOT + len(CQ_ROWS) + 2
label(cq, FOOT, 2, "Payment terms: 14 days from delivery, bank transfer.", size=10)
label(cq, FOOT + 1, 2, "Recurring items are billed monthly for the length of the contract.", size=10)
label(cq, FOOT + 2, 2, "This quotation is valid until the date shown above.", size=10, color=MUTED)

for col, width in (("A", 2.5), ("B", 5), ("C", 42), ("D", 11), ("E", 8), ("F", 13), ("G", 15)):
    cq.column_dimensions[col].width = width
cq.print_area = f"A1:G{FOOT + 2}"
cq.protection.sheet = True
cq.protection.enable()

wb.save(OUT / "quotation.xlsx")

# ---------------- independent check (the workbook itself uses formulas) ----------------
one_off = sum(round(int(l["quantity"]) * price_of[l["product"]] * (1 - float(l["discount"])), 2)
              for l in request if billing_of[l["product"]] == "one-off")
monthly = sum(round(int(l["quantity"]) * price_of[l["product"]] * (1 - float(l["discount"])), 2)
              for l in request if billing_of[l["product"]] == "monthly")
cost = sum(int(l["quantity"]) * cost_of[l["product"]] * (TERM_MONTHS if billing_of[l["product"]] == "monthly" else 1)
           for l in request)
contract_net = one_off + monthly * TERM_MONTHS
margin_pct = (contract_net - cost) / contract_net

print(f"OK -> output/quotation.xlsx (formulas recalculate when opened in Excel)")
print(f"   {len(request)} line items from a {len(catalogue)}-product catalogue, {TERM_MONTHS}-month term")
print(f"   upfront {one_off:,.2f} zl + {monthly:,.2f} zl/month -> contract net {contract_net:,.2f} zl")
print(f"   gross margin {contract_net - cost:,.2f} zl ({margin_pct:.1%})")
